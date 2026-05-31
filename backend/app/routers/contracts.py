"""
Contracts API Router - Contract Registry
"""
import logging
from typing import List, Optional, Dict
import uuid
import json
from datetime import date, datetime, timedelta
from calendar import monthrange

logger = logging.getLogger(__name__)

import io
import zipfile

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form, Response, Request
from fastapi.responses import StreamingResponse
from sqlalchemy import select, or_, cast, String, delete
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm.attributes import set_committed_value
from pydantic import BaseModel
import httpx
from urllib.parse import quote

from app.database.session import get_db
from app.core.auth_middleware import CurrentUser
from app.models import (
    Contract,
    Deal,
    SubcontractorCard,
    ContractDocument,
    ContractDocumentProductLink,
    DealProduct,
    IncomeExpenseEntry,
    TreasuryAllocation,
    TreasuryTransaction,
    Company,
    Stage,
    SubcontractorStage,
    Product,
)
from app.services.permissions import allowed_deal_ids, get_section_permissions, ensure_can_edit_record
from app.schemas.contract import ContractCreate, ContractUpdate, ContractResponse
from app.schemas.contract_document import ContractDocumentUpdate, ContractDocumentResponse
from app.schemas.contract_card import ContractCardResponse, ContractPaymentSummary, ContractStageSummary
from app.schemas.income_expense import IncomeExpenseEntryResponse, PaymentHistoryItem
from app.services.data_health import safe_refresh_deal_health_issues, safe_refresh_orphan_health_issues
from app.services.event_log import log_event
from app.services.sequence_lock import sequence_lock
import mimetypes
from app.services.storage import (
    clean_name,
    ensure_path,
    upload_bytes_with_safe_extension,
    get_download_href,
    read_file_bytes,
    delete_path,
    storage_available,
    is_local_storage,
)
from app.core.config import settings

router = APIRouter()

ALLOWED_CONTRACT_TYPES = {"general_contractor", "subcontractor", "services", "labor"}
ALLOWED_DOC_TYPES = {"contract", "addendum", "act", "invoice"}
ALLOWED_DOC_STATUSES = {"draft", "signing", "signed", "canceled"}
ALLOWED_FILE_KINDS = {"pdf", "edit"}
ALLOWED_EXPENSE_FREQUENCIES = {"week", "month", "quarter", "year"}


def _user_display_name(user) -> Optional[str]:
    """Best-effort display name snapshot for upload audit.

    Prefers full_name, falls back to username/email/id. Stored as a string
    snapshot (not a FK) so the attribution survives user renames/deletes.
    """
    if user is None:
        return None
    for attr in ("full_name", "name", "username", "email"):
        value = getattr(user, attr, None)
        if value:
            return str(value)
    user_id = getattr(user, "id", None)
    return str(user_id) if user_id else None


def _parse_uuid(value: Optional[str]):
    if not value:
        return None
    try:
        return value if isinstance(value, uuid.UUID) else uuid.UUID(str(value))
    except (ValueError, TypeError):
        return None


def _id_conditions(column, value):
    variants = []
    try:
        parsed = value if isinstance(value, uuid.UUID) else uuid.UUID(str(value))
        variants.extend([parsed, str(parsed), parsed.hex])
    except (ValueError, TypeError):
        variants.append(str(value))
    return or_(*[column == v for v in variants])


def _status_from_paid(amount: float, paid_amount: float) -> str:
    if amount <= 0:
        return "paid"
    if paid_amount <= 0:
        return "unpaid"
    if paid_amount + 1e-9 < amount:
        return "partial"
    return "paid"


def _normalize_contract_key(value: Optional[str]) -> str:
    return str(value or "").replace("-", "").lower()


def _contract_snapshot(contract: Optional[Contract]) -> Dict[str, Optional[str]]:
    if not contract:
        return {}
    return {
        "contract_number": contract.contract_number,
        "status": contract.status,
        "contract_type": contract.contract_type,
        "deal_id": str(contract.deal_id) if getattr(contract, "deal_id", None) else None,
        "subcontractor_card_id": str(contract.subcontractor_card_id) if getattr(contract, "subcontractor_card_id", None) else None,
        "contract_date": contract.contract_date.isoformat() if getattr(contract, "contract_date", None) else None,
        "amount": str(contract.amount) if getattr(contract, "amount", None) is not None else None,
    }


def _contract_changes(before: Dict[str, Optional[str]], after: Dict[str, Optional[str]]) -> Dict[str, Dict[str, Optional[str]]]:
    changes: Dict[str, Dict[str, Optional[str]]] = {}
    for key in set(before.keys()) | set(after.keys()):
        if before.get(key) != after.get(key):
            changes[key] = {"before": before.get(key), "after": after.get(key)}
    return changes


async def _compute_service_paid_amounts(
    db: AsyncSession,
    contracts: List[Contract],
) -> Dict[str, float]:
    """Return paid sums by contract_id for service contracts only.

    Paid sum is calculated from actual treasury allocations linked to income/expense
    entries of a contract (not from planned amounts).
    """
    service_contracts = [
        c for c in contracts
        if c and getattr(c, "id", None) and getattr(c, "contract_type", None) == "services"
    ]
    if not service_contracts:
        return {}

    variants: set[str] = set()
    normalized_to_canonical: Dict[str, str] = {}
    for contract in service_contracts:
        canonical_id = str(contract.id)
        parsed = _parse_uuid(contract.id)
        contract_variants = [canonical_id]
        if parsed:
            contract_variants.extend([str(parsed), parsed.hex])
        else:
            contract_variants.append(canonical_id.replace("-", ""))
        for variant in contract_variants:
            if not variant:
                continue
            variants.add(str(variant))
            normalized_to_canonical[_normalize_contract_key(variant)] = canonical_id

    if not variants:
        return {}

    result = await db.execute(
        select(IncomeExpenseEntry.contract_id, TreasuryAllocation.amount)
        .join(TreasuryAllocation, TreasuryAllocation.income_expense_id == IncomeExpenseEntry.id)
        .where(IncomeExpenseEntry.contract_id.in_(list(variants)))
    )

    paid_map: Dict[str, float] = {}
    for contract_id, amount in result.all():
        canonical_id = normalized_to_canonical.get(_normalize_contract_key(contract_id))
        if not canonical_id:
            continue
        paid_map[canonical_id] = paid_map.get(canonical_id, 0.0) + float(amount or 0.0)

    return paid_map


async def _apply_service_paid_amounts(
    db: AsyncSession,
    contracts: List[Contract],
) -> None:
    """Set transient `amount` for service contracts to paid allocations sum."""
    if not contracts:
        return
    paid_map = await _compute_service_paid_amounts(db, contracts)
    for contract in contracts:
        if getattr(contract, "contract_type", None) != "services":
            continue
        value = float(paid_map.get(str(contract.id), 0.0))
        set_committed_value(contract, "amount", value)


def _add_months(value: date, months: int) -> date:
    year = value.year + (value.month - 1 + months) // 12
    month = (value.month - 1 + months) % 12 + 1
    day = min(value.day, monthrange(year, month)[1])
    return date(year, month, day)


def _add_months_with_anchor(value: date, months: int, anchor_day: int, anchor_is_eom: bool) -> date:
    year = value.year + (value.month - 1 + months) // 12
    month = (value.month - 1 + months) % 12 + 1
    last_day = monthrange(year, month)[1]
    day = last_day if anchor_is_eom else min(anchor_day, last_day)
    return date(year, month, day)


class ContractExpenseCreate(BaseModel):
    amount: float
    plan_date: date
    actual_date: Optional[date] = None
    category_code: Optional[str] = None


class ContractExpenseUpdate(BaseModel):
    amount: Optional[float] = None
    plan_date: Optional[date] = None
    actual_date: Optional[date] = None
    category_code: Optional[str] = None


class ContractExpenseBulkCreate(BaseModel):
    category_code: Optional[str] = None
    amount: float
    start_date: date
    frequency: str
    periods: int


class DocumentProductLinksUpdate(BaseModel):
    deal_product_ids: List[str] = []


class ContractExpenseCardResponse(BaseModel):
    summary: ContractPaymentSummary
    items: List[IncomeExpenseEntryResponse]


async def _ensure_single_contract(
    db: AsyncSession,
    contract_type: Optional[str],
    deal_id: Optional[str],
    subcontractor_card_id: Optional[str],
    exclude_contract_id: Optional[str] = None,
) -> None:
    if contract_type == "general_contractor" and deal_id:
        existing = await Contract.get_by_deal_id(db, deal_id)
        for item in existing:
            if item.contract_type == "general_contractor":
                if not exclude_contract_id or str(item.id) != str(exclude_contract_id):
                    raise HTTPException(
                        status_code=400,
                        detail="Deal already has a general contractor contract",
                    )


async def _get_or_create_subcontractor_card(
    db: AsyncSession,
    executor_id: Optional[str],
    deal_id: Optional[str] = None,
):
    if not executor_id:
        return None
    result = await db.execute(
        select(SubcontractorCard).where(_id_conditions(SubcontractorCard.company_id, executor_id))
    )
    card = result.scalar_one_or_none()
    if card:
        return card
    company = await Company.get_by_id(db, executor_id)
    payload = {
        "title": company.name if company and getattr(company, "name", None) else "Субподрядчик",
        "company_id": str(executor_id),
    }
    if deal_id:
        deal = await Deal.get_by_id(db, deal_id)
        if deal:
            payload.update(
                {
                    "obj_name": deal.obj_name,
                    "address": deal.address,
                    "object_type": deal.object_type,
                    "object_area": deal.object_area,
                    "customer_id": deal.customer_id,
                    "general_contractor_id": deal.our_company_id,
                }
            )
    return await SubcontractorCard.create(db, **payload)


@router.get("/")
async def get_contracts(
    request: Request,
    skip: int = 0,
    limit: int = 100,
    search: Optional[str] = Query(None),
    contract_type: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    customer_id: Optional[str] = Query(None),
    executor_id: Optional[str] = Query(None),
    deal_id: Optional[str] = Query(None),
    subcontractor_card_id: Optional[str] = Query(None),
    sort_by: Optional[str] = Query("contract_date"),
    sort_dir: Optional[str] = Query("desc"),
    db: AsyncSession = Depends(get_db),
    user=Depends(CurrentUser),
):
    """List contracts with pagination, search and filters"""
    try:
        read_all, read_assigned = await get_section_permissions(db, user.role_id, "contracts")
        allowed = None
        if not read_all:
            if not read_assigned:
                return {"items": [], "total": 0, "stats": {"total": 0, "approval": 0, "in_progress": 0, "completed": 0}}
            allowed = await allowed_deal_ids(db, request, user)
            if allowed == []:
                return {"items": [], "total": 0, "stats": {"total": 0, "approval": 0, "in_progress": 0, "completed": 0}}

        # Legacy behaviour: filter by deal_id or subcontractor_card_id returns plain list
        if deal_id and not search and not contract_type and not status:
            contracts = await Contract.get_by_deal_id(db, deal_id)
            if allowed is not None:
                allowed_set = set(allowed)
                contracts = [c for c in contracts if c.deal_id and str(c.deal_id) in allowed_set]
            await _apply_service_paid_amounts(db, contracts)
            return contracts
        if subcontractor_card_id:
            contracts = await Contract.get_by_subcontractor_card_id(db, subcontractor_card_id)
            if allowed is not None:
                allowed_set = set(allowed)
                contracts = [c for c in contracts if c.deal_id and str(c.deal_id) in allowed_set]
            await _apply_service_paid_amounts(db, contracts)
            return contracts

        # New behaviour: search with pagination
        items, total = await Contract.search_all(
            db,
            skip=skip,
            limit=limit,
            search=search,
            contract_type=contract_type,
            status=status,
            customer_id=customer_id,
            executor_id=executor_id,
            deal_id=deal_id,
            sort_by=sort_by,
            sort_dir=sort_dir,
            allowed_deal_ids=allowed,
        )
        await _apply_service_paid_amounts(db, items)

        # Stats by status
        stats_by_status = await Contract.count_by_status(db, allowed_deal_ids=allowed)
        total_all = sum(stats_by_status.values())

        return {
            "items": items,
            "total": total,
            "stats": {
                "total": total_all,
                "approval": stats_by_status.get("approval", 0),
                "in_progress": stats_by_status.get("in_progress", 0),
                "completed": stats_by_status.get("completed", 0),
            }
        }
    except Exception as e:
        print(f"Error getting contracts: {e}")
        import traceback
        traceback.print_exc()
        return {"items": [], "total": 0, "stats": {"total": 0, "approval": 0, "in_progress": 0, "completed": 0}}


@router.get("/deal/{deal_id}", response_model=List[ContractResponse])
async def get_contracts_by_deal(
    deal_id: str,
    request: Request,
    db: AsyncSession = Depends(get_db),
    user=Depends(CurrentUser),
):
    """List contracts linked to a deal"""
    read_all, read_assigned = await get_section_permissions(db, user.role_id, "contracts")
    if not read_all:
        if not read_assigned:
            return []
        allowed = await allowed_deal_ids(db, request, user)
        if allowed is not None and str(deal_id) not in set(allowed):
            return []
    contracts = await Contract.get_by_deal_id(db, deal_id)
    await _apply_service_paid_amounts(db, contracts)
    return contracts


@router.get("/subcontractor/{subcontractor_id}", response_model=List[ContractResponse])
async def get_contracts_by_subcontractor(
    subcontractor_id: str,
    request: Request,
    db: AsyncSession = Depends(get_db),
    user=Depends(CurrentUser),
):
    """List contracts linked to a subcontractor card"""
    read_all, read_assigned = await get_section_permissions(db, user.role_id, "contracts")
    if not read_all:
        if not read_assigned:
            return []
        allowed = await allowed_deal_ids(db, request, user)
    else:
        allowed = None
    contracts = await Contract.get_by_subcontractor_card_id(db, subcontractor_id)
    if allowed is not None:
        allowed_set = set(allowed)
        contracts = [c for c in contracts if c.deal_id and str(c.deal_id) in allowed_set]
    await _apply_service_paid_amounts(db, contracts)
    return contracts


@router.post("/", response_model=ContractResponse)
async def create_contract(
    contract: ContractCreate,
    db: AsyncSession = Depends(get_db),
    user=Depends(CurrentUser),
):
    """Create a contract"""
    if contract.contract_type not in ALLOWED_CONTRACT_TYPES:
        raise HTTPException(status_code=400, detail="Unsupported contract_type")
    if (
        contract.contract_type == "general_contractor"
        and contract.deal_id
        and contract.subcontractor_card_id
    ):
        raise HTTPException(status_code=400, detail="Contract must be linked to deal or subcontractor, not both")
    if contract.deal_id:
        deal = await Deal.get_by_id(db, contract.deal_id)
        if not deal:
            raise HTTPException(status_code=404, detail="Deal not found")
    if contract.subcontractor_card_id:
        card = await SubcontractorCard.get_by_id(db, contract.subcontractor_card_id)
        if not card:
            raise HTTPException(status_code=404, detail="Subcontractor not found")
    create_payload = contract.dict()
    if create_payload.get("contract_type") == "services":
        create_payload["amount"] = 0.0
    if contract.contract_type == "subcontractor" and not create_payload.get("subcontractor_card_id"):
        card = await _get_or_create_subcontractor_card(
            db,
            create_payload.get("executor_id"),
            create_payload.get("deal_id"),
        )
        if card:
            create_payload["subcontractor_card_id"] = _parse_uuid(card.id) or str(card.id)
    await _ensure_single_contract(
        db,
        contract.contract_type,
        str(create_payload.get("deal_id")) if create_payload.get("deal_id") else None,
        str(create_payload.get("subcontractor_card_id")) if create_payload.get("subcontractor_card_id") else None,
    )
    db_contract = await Contract.create(db, **create_payload)
    await _apply_service_paid_amounts(db, [db_contract])
    if db_contract.deal_id:
        try:
            await log_event(
                db,
                entity_type="deal",
                entity_id=str(db_contract.deal_id),
                action="contract.create",
                created_by=str(getattr(user, "id", "")),
                details={
                    "deal_id": str(db_contract.deal_id),
                    "contract_id": str(db_contract.id),
                    "contract_number": db_contract.contract_number,
                    "status": db_contract.status,
                    "contract_type": db_contract.contract_type,
                    "amount": float(db_contract.amount or 0),
                },
            )
        except Exception:
            pass
    await safe_refresh_deal_health_issues(db, db_contract.deal_id)
    if not db_contract.deal_id:
        await safe_refresh_orphan_health_issues(db)
    # Event bus: новый договор привязан к проекту — внешние подписчики
    # (BI, Telegram-канал договорного отдела) узнают через outbox.
    from app.services.event_outbox import emit_event_safe
    await emit_event_safe(
        db,
        event_type="contract.after_create",
        entity_type="contract",
        entity_id=str(db_contract.id),
        payload={
            "id": str(db_contract.id),
            "contract_number": db_contract.contract_number,
            "contract_type": db_contract.contract_type,
            "status": db_contract.status,
            "amount": float(db_contract.amount or 0),
            "deal_id": str(db_contract.deal_id) if db_contract.deal_id else None,
            "subcontractor_card_id": str(db_contract.subcontractor_card_id) if db_contract.subcontractor_card_id else None,
            "created_by_user_id": str(getattr(user, "id", "")) or None,
        },
    )
    await db.commit()
    return db_contract


@router.get("/{contract_id}", response_model=ContractResponse)
async def get_contract(
    contract_id: str,
    request: Request,
    db: AsyncSession = Depends(get_db),
    user=Depends(CurrentUser),
):
    """Get a contract by ID"""
    contract = await Contract.get_by_id(db, contract_id)
    if not contract:
        raise HTTPException(status_code=404, detail="Contract not found")
    read_all, read_assigned = await get_section_permissions(db, user.role_id, "contracts")
    if not read_all:
        if not read_assigned:
            raise HTTPException(status_code=404, detail="Contract not found")
        allowed = await allowed_deal_ids(db, request, user)
        if allowed is not None:
            if not contract.deal_id or str(contract.deal_id) not in set(allowed):
                raise HTTPException(status_code=404, detail="Contract not found")
    await _apply_service_paid_amounts(db, [contract])
    return contract


@router.put("/{contract_id}", response_model=ContractResponse)
async def update_contract(
    contract_id: str,
    contract_update: ContractUpdate,
    request: Request,
    db: AsyncSession = Depends(get_db),
    user=Depends(CurrentUser),
):
    """Update a contract"""
    if contract_update.contract_type and contract_update.contract_type not in ALLOWED_CONTRACT_TYPES:
        raise HTTPException(status_code=400, detail="Unsupported contract_type")
    if contract_update.deal_id:
        deal = await Deal.get_by_id(db, contract_update.deal_id)
        if not deal:
            raise HTTPException(status_code=404, detail="Deal not found")
    if contract_update.subcontractor_card_id:
        card = await SubcontractorCard.get_by_id(db, contract_update.subcontractor_card_id)
        if not card:
            raise HTTPException(status_code=404, detail="Subcontractor not found")
    existing = await Contract.get_by_id(db, contract_id)
    if not existing:
        raise HTTPException(status_code=404, detail="Contract not found")
    await ensure_can_edit_record(db, request, user, "contracts", existing)
    before_snapshot = _contract_snapshot(existing)
    update_payload = contract_update.dict(exclude_unset=True)
    next_type = contract_update.contract_type or existing.contract_type
    if "deal_id" in update_payload:
        next_deal_id = str(update_payload["deal_id"]) if update_payload["deal_id"] else None
    else:
        next_deal_id = str(existing.deal_id) if existing.deal_id else None
    if "subcontractor_card_id" in update_payload:
        next_sub_id = str(update_payload["subcontractor_card_id"]) if update_payload["subcontractor_card_id"] else None
    else:
        next_sub_id = str(existing.subcontractor_card_id) if existing.subcontractor_card_id else None
    if next_deal_id and next_sub_id and next_type != "subcontractor":
        raise HTTPException(status_code=400, detail="Contract must be linked to deal or subcontractor, not both")
    executor_id = update_payload.get("executor_id") if "executor_id" in update_payload else existing.executor_id
    if next_type == "subcontractor" and not next_sub_id and executor_id:
        card = await _get_or_create_subcontractor_card(db, executor_id, next_deal_id)
        if card:
            next_sub_id = str(card.id)
            update_payload["subcontractor_card_id"] = _parse_uuid(card.id) or str(card.id)
    if next_type == "services":
        update_payload["amount"] = 0.0
    await _ensure_single_contract(
        db,
        next_type,
        next_deal_id,
        next_sub_id,
        exclude_contract_id=contract_id,
    )

    # ── Event Bus v2: before_status_change ──────────────────────────
    # Если в обновлении меняется статус — даём in-process обработчикам
    # шанс отменить или модифицировать payload. Cancel → 400 в ответ
    # пользователю; Mutate → правки уже применены к ctx.payload, тут
    # их синхронизируем обратно в update_payload (для статус-перехода
    # на практике редко, но контракт держим).
    from app.services.event_dispatcher import (
        Cancel as _EBCancel,
        EventContext as _EBCtx,
        Mutate as _EBMutate,
        dispatch_before,
        dispatch_after,
    )
    status_after = update_payload.get("status")
    status_before = existing.status
    status_change_ctx = None
    if status_after and status_after != status_before:
        status_change_ctx = _EBCtx(
            event_key="contract.before_status_change",
            entity_type="contract",
            entity_id=str(contract_id),
            payload={
                "status_before": status_before,
                "status_after": status_after,
                "amount": float(existing.amount or 0),
                "paid_amount": float(existing.paid_amount or 0),
                "contract_type": existing.contract_type,
                "deal_id": str(existing.deal_id) if existing.deal_id else None,
            },
            user_id=str(getattr(user, "id", "")) or None,
        )
        result = await dispatch_before("contract.before_status_change", status_change_ctx)
        if isinstance(result, _EBCancel):
            raise HTTPException(status_code=400, detail=result.reason)
        if isinstance(result, _EBMutate):
            # Применяем только те поля, которые реально входят в update-схему.
            for key in ("status",):
                if key in result.payload_patch:
                    update_payload[key] = result.payload_patch[key]

    contract = await Contract.update(db, contract_id, **update_payload)
    if not contract:
        raise HTTPException(status_code=404, detail="Contract not found")
    await _apply_service_paid_amounts(db, [contract])

    # ── Event Bus v2: after_status_change ───────────────────────────
    # Outbox-запись + in-process after-хендлеры. Ошибка тут не валит
    # запрос, бизнес-данные уже зафиксированы.
    if status_change_ctx is not None:
        status_change_ctx.event_key = "contract.after_status_change"
        # обновим snapshot, статус мог быть мутирован
        status_change_ctx.payload["status_after"] = contract.status
        try:
            await dispatch_after(db, "contract.after_status_change", status_change_ctx)
        except Exception:
            logger.exception("dispatch_after contract.after_status_change failed")
    if contract.deal_id:
        try:
            await log_event(
                db,
                entity_type="deal",
                entity_id=str(contract.deal_id),
                action="contract.update",
                created_by=str(getattr(user, "id", "")),
                details={
                    "deal_id": str(contract.deal_id),
                    "contract_id": str(contract.id),
                    "contract_number": contract.contract_number,
                    "status": contract.status,
                    "contract_type": contract.contract_type,
                    "changes": _contract_changes(before_snapshot, _contract_snapshot(contract)),
                },
            )
        except Exception:
            pass
    await safe_refresh_deal_health_issues(db, contract.deal_id)
    if before_snapshot.get("deal_id") and before_snapshot.get("deal_id") != str(contract.deal_id or ""):
        await safe_refresh_deal_health_issues(db, before_snapshot.get("deal_id"))
    if not contract.deal_id:
        await safe_refresh_orphan_health_issues(db)

    # Event Bus v2: общая after_update для внешних подписчиков.
    # status_change emit-ится отдельно выше (если был); это — общее
    # «контракт обновлён», часто подписываются для синхронизации в 1С.
    from app.services.event_outbox import emit_event_safe
    await emit_event_safe(
        db,
        event_type="contract.after_update",
        entity_type="contract",
        entity_id=str(contract.id),
        payload={
            "id": str(contract.id),
            "deal_id": str(contract.deal_id) if contract.deal_id else None,
            "status": contract.status,
            "contract_type": contract.contract_type,
            "amount": float(contract.amount or 0),
            "paid_amount": float(contract.paid_amount or 0),
            "changes": _contract_changes(before_snapshot, _contract_snapshot(contract)),
        },
    )
    return contract


@router.patch("/{contract_id}/link")
async def link_contract(
    contract_id: str,
    request: Request,
    deal_id: Optional[str] = None,
    subcontractor_card_id: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    user=Depends(CurrentUser),
):
    """Link/unlink contract to deal or subcontractor"""
    contract = await Contract.get_by_id(db, contract_id)
    if not contract:
        raise HTTPException(status_code=404, detail="Contract not found")
    await ensure_can_edit_record(db, request, user, "contracts", contract)
    previous_deal_id = str(contract.deal_id) if contract.deal_id else None
    if deal_id and subcontractor_card_id and contract.contract_type != "subcontractor":
        raise HTTPException(status_code=400, detail="Contract must be linked to deal or subcontractor, not both")
    if deal_id:
        deal = await Deal.get_by_id(db, deal_id)
        if not deal:
            raise HTTPException(status_code=404, detail="Deal not found")
    if subcontractor_card_id:
        card = await SubcontractorCard.get_by_id(db, subcontractor_card_id)
        if not card:
            raise HTTPException(status_code=404, detail="Subcontractor not found")
    if contract.contract_type == "subcontractor" and not subcontractor_card_id and not contract.subcontractor_card_id:
        card = await _get_or_create_subcontractor_card(db, contract.executor_id, deal_id)
        if card:
            subcontractor_card_id = str(card.id)
    await _ensure_single_contract(
        db,
        contract.contract_type,
        deal_id,
        subcontractor_card_id,
        exclude_contract_id=contract_id,
    )
    contract = await Contract.update(
        db,
        contract_id,
        deal_id=deal_id,
        subcontractor_card_id=subcontractor_card_id
    )
    if not contract:
        raise HTTPException(status_code=404, detail="Contract not found")
    target_deal_id = str(contract.deal_id or deal_id or previous_deal_id or "")
    if target_deal_id:
        try:
            await log_event(
                db,
                entity_type="deal",
                entity_id=target_deal_id,
                action="contract.link",
                created_by=str(getattr(user, "id", "")),
                details={
                    "deal_id": target_deal_id,
                    "contract_id": str(contract.id),
                    "contract_number": contract.contract_number,
                    "status": contract.status,
                    "changes": {
                        "deal_id": {"before": previous_deal_id, "after": str(contract.deal_id) if contract.deal_id else None},
                        "subcontractor_card_id": {"before": None, "after": str(contract.subcontractor_card_id) if contract.subcontractor_card_id else None},
                    },
                },
            )
        except Exception:
            pass
    await safe_refresh_deal_health_issues(db, contract.deal_id)
    if previous_deal_id and previous_deal_id != str(contract.deal_id or ""):
        await safe_refresh_deal_health_issues(db, previous_deal_id)
    if not contract.deal_id:
        await safe_refresh_orphan_health_issues(db)
    return {"message": "Link updated"}


@router.delete("/{contract_id}")
async def delete_contract(
    contract_id: str,
    request: Request,
    db: AsyncSession = Depends(get_db),
    user=Depends(CurrentUser),
):
    """Delete a contract"""
    contract = await Contract.get_by_id(db, contract_id)
    if not contract:
        raise HTTPException(status_code=404, detail="Contract not found")
    await ensure_can_edit_record(db, request, user, "contracts", contract)
    success = await Contract.delete(db, contract_id)
    if not success:
        raise HTTPException(status_code=404, detail="Contract not found")
    if contract.deal_id:
        try:
            await log_event(
                db,
                entity_type="deal",
                entity_id=str(contract.deal_id),
                action="contract.delete",
                created_by=str(getattr(user, "id", "")),
                details={
                    "deal_id": str(contract.deal_id),
                    "contract_id": str(contract.id),
                    "contract_number": contract.contract_number,
                    "status": contract.status,
                },
            )
        except Exception:
            pass
    await safe_refresh_deal_health_issues(db, contract.deal_id)
    if not contract.deal_id:
        await safe_refresh_orphan_health_issues(db)

    # Event Bus v2: после удаления — отправляем outbox-событие, чтобы
    # внешний consumer (1С, Диадок и т.п.) мог зачистить связанные
    # записи. id всё ещё знаем — модель уже подгружена выше.
    from app.services.event_outbox import emit_event_safe
    await emit_event_safe(
        db,
        event_type="contract.after_delete",
        entity_type="contract",
        entity_id=str(contract.id),
        payload={
            "id": str(contract.id),
            "deal_id": str(contract.deal_id) if contract.deal_id else None,
            "contract_number": contract.contract_number,
            "status": contract.status,
        },
    )
    return {"message": "Contract deleted"}


async def _build_payment_responses(
    db: AsyncSession,
    entries: List[IncomeExpenseEntry],
) -> List[IncomeExpenseEntryResponse]:
    company_ids = {e.payer_id for e in entries if e.payer_id} | {e.payee_id for e in entries if e.payee_id}
    deal_ids = {e.deal_id for e in entries if e.deal_id}
    contract_ids = {e.contract_id for e in entries if e.contract_id}

    company_map: Dict[str, Company] = {}
    if company_ids:
        company_result = await db.execute(select(Company).where(Company.id.in_(company_ids)))
        company_map = {c.id: c for c in company_result.scalars().all()}

    deal_map: Dict[str, Deal] = {}
    if deal_ids:
        deal_result = await db.execute(select(Deal).where(Deal.id.in_(deal_ids)))
        deal_map = {d.id: d for d in deal_result.scalars().all()}

    contract_map: Dict[str, Contract] = {}
    if contract_ids:
        contract_uuid_ids = {cid for cid in (_parse_uuid(value) for value in contract_ids) if cid}
        if contract_uuid_ids:
            contract_result = await db.execute(select(Contract).where(Contract.id.in_(contract_uuid_ids)))
            for contract in contract_result.scalars().all():
                contract_map[str(contract.id)] = contract
                contract_map[contract.id.hex] = contract

    history_map: Dict[str, List[PaymentHistoryItem]] = {}
    paid_map: Dict[str, float] = {}
    entry_ids = [e.id for e in entries]
    if entry_ids:
        alloc_result = await db.execute(
            select(TreasuryAllocation, TreasuryTransaction)
            .join(TreasuryTransaction, TreasuryAllocation.transaction_id == TreasuryTransaction.id)
            .where(TreasuryAllocation.income_expense_id.in_(entry_ids))
            .order_by(TreasuryTransaction.transaction_date.asc())
        )
        for allocation, tx in alloc_result.all():
            entry_id = allocation.income_expense_id
            if not entry_id:
                continue
            history_map.setdefault(entry_id, []).append(PaymentHistoryItem(
                transaction_id=str(tx.id),
                transaction_date=tx.transaction_date,
                amount=allocation.amount,
                doc_num=tx.doc_num,
                allocation_id=str(allocation.id),
                category_code=allocation.category_code
            ))
            paid_map[entry_id] = paid_map.get(entry_id, 0.0) + allocation.amount

    responses: List[IncomeExpenseEntryResponse] = []
    for entry in entries:
        paid_amount = paid_map.get(entry.id, 0.0)
        payment_status = _status_from_paid(entry.amount, paid_amount)
        contract = None
        if entry.contract_id:
            contract = contract_map.get(entry.contract_id)
            if not contract:
                parsed_id = _parse_uuid(entry.contract_id)
                if parsed_id:
                    contract = contract_map.get(str(parsed_id)) or contract_map.get(parsed_id.hex)
        responses.append(IncomeExpenseEntryResponse(
            id=str(entry.id),
            direction=entry.direction,
            amount=entry.amount,
            plan_date=entry.plan_date,
            actual_date=entry.actual_date,
            payer_id=entry.payer_id,
            payee_id=entry.payee_id,
            deal_id=entry.deal_id,
            contract_id=entry.contract_id,
            stage_id=entry.stage_id,
            category_code=entry.category_code,
            payer_name=company_map.get(entry.payer_id).name if entry.payer_id in company_map else None,
            payee_name=company_map.get(entry.payee_id).name if entry.payee_id in company_map else None,
            deal_title=deal_map.get(entry.deal_id).title if entry.deal_id in deal_map else None,
            contract_number=contract.contract_number if contract else None,
            payment_status=payment_status,
            paid_amount=paid_amount,
            payments_history=history_map.get(entry.id, []),
            warning=None,
        ))

    return responses


async def _get_service_contract_or_400(db: AsyncSession, contract_id: str) -> Contract:
    contract = await Contract.get_by_id(db, contract_id)
    if not contract:
        raise HTTPException(status_code=404, detail="Contract not found")
    if contract.contract_type != "services":
        raise HTTPException(status_code=400, detail="Expense card available только для договоров оказания услуг")
    return contract


def _contract_id_conditions(contract_id: str):
    parsed = _parse_uuid(contract_id)
    if not parsed:
        return or_(IncomeExpenseEntry.contract_id == str(contract_id))
    return or_(
        IncomeExpenseEntry.contract_id == str(parsed),
        IncomeExpenseEntry.contract_id == parsed.hex,
    )


def _next_period_date(value: date, frequency: str) -> date:
    if frequency == "week":
        return value + timedelta(days=7)
    if frequency == "month":
        return _add_months(value, 1)
    if frequency == "quarter":
        return _add_months(value, 3)
    if frequency == "year":
        return _add_months(value, 12)
    return value


async def _build_stage_payment_maps(
    db: AsyncSession,
    stage_ids: List[str],
) -> Dict[str, Dict[str, float]]:
    if not stage_ids:
        return {"total": {}, "paid": {}}
    stage_ids_set = set(stage_ids)
    stage_ids_set.update({sid.replace("-", "") for sid in stage_ids})
    result = await db.execute(
        select(IncomeExpenseEntry).where(IncomeExpenseEntry.stage_id.in_(list(stage_ids_set)))
    )
    entries = result.scalars().all()
    entry_ids = [e.id for e in entries]
    total_map: Dict[str, float] = {}
    paid_map: Dict[str, float] = {}
    for entry in entries:
        if entry.stage_id:
            total_map[entry.stage_id] = total_map.get(entry.stage_id, 0.0) + entry.amount
    if entry_ids:
        alloc_result = await db.execute(
            select(TreasuryAllocation, TreasuryTransaction)
            .join(TreasuryTransaction, TreasuryAllocation.transaction_id == TreasuryTransaction.id)
            .where(TreasuryAllocation.income_expense_id.in_(entry_ids))
        )
        for allocation, tx in alloc_result.all():
            entry_id = allocation.income_expense_id
            if not entry_id:
                continue
            entry = next((e for e in entries if e.id == entry_id), None)
            if not entry or not entry.stage_id:
                continue
            paid_map[entry.stage_id] = paid_map.get(entry.stage_id, 0.0) + allocation.amount
    return {"total": total_map, "paid": paid_map}


async def _renumber_documents(db: AsyncSession, contract_id: str, doc_type: str) -> None:
    documents = await ContractDocument.get_by_contract_and_type(db, contract_id, doc_type)
    for idx, doc in enumerate(documents, start=1):
        if doc.number_in_contract != idx:
            await ContractDocument.update(db, str(doc.id), number_in_contract=idx)


def _document_amount(value: Optional[float]) -> Optional[float]:
    if value in (None, ""):
        return None
    try:
        amount = float(value)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="Invalid amount")
    if amount < 0:
        raise HTTPException(status_code=400, detail="Amount must be positive")
    return amount


def _parse_product_ids(raw: Optional[str]) -> List[str]:
    if not raw:
        return []
    if isinstance(raw, list):
        source = raw
    else:
        text = str(raw).strip()
        if not text:
            return []
        try:
            parsed = json.loads(text)
            source = parsed if isinstance(parsed, list) else [parsed]
        except json.JSONDecodeError:
            source = text.split(",")
    result: List[str] = []
    seen = set()
    for value in source:
        item = str(value or "").strip()
        if not item:
            continue
        key = item.replace("-", "").lower()
        if key in seen:
            continue
        seen.add(key)
        result.append(item)
    return result


def _normalize_link_key(value: Optional[str]) -> str:
    return str(value or "").replace("-", "").lower()


def _deal_product_name(deal_product: DealProduct, product: Optional[Product] = None) -> str:
    custom_name = getattr(deal_product, "custom_name", None)
    if custom_name:
        return custom_name
    if product and getattr(product, "name", None):
        return product.name
    loaded_product = getattr(deal_product, "product", None)
    if loaded_product and getattr(loaded_product, "name", None):
        return loaded_product.name
    return "Товар"


async def _attach_document_product_links(db: AsyncSession, documents: List[ContractDocument]) -> List[ContractDocument]:
    if not documents:
        return documents
    doc_ids = [str(doc.id) for doc in documents if getattr(doc, "id", None)]
    for doc in documents:
        setattr(doc, "linked_products", [])
    if not doc_ids:
        return documents

    result = await db.execute(
        select(ContractDocumentProductLink, DealProduct, Product)
        .join(DealProduct, ContractDocumentProductLink.deal_product_id == DealProduct.id)
        .outerjoin(Product, DealProduct.product_id == Product.id)
        .where(ContractDocumentProductLink.contract_document_id.in_(doc_ids))
    )
    by_doc: Dict[str, List[Dict[str, Optional[str]]]] = {}
    for link, deal_product, product in result.all():
        key = str(link.contract_document_id)
        by_doc.setdefault(key, []).append({
            "deal_product_id": str(deal_product.id),
            "product_name": _deal_product_name(deal_product, product),
            "custom_name": getattr(deal_product, "custom_name", None),
        })
    for doc in documents:
        setattr(doc, "linked_products", by_doc.get(str(doc.id), []))
    return documents


async def _set_invoice_product_links(
    db: AsyncSession,
    document: ContractDocument,
    deal_product_ids: List[str],
) -> None:
    await db.execute(
        delete(ContractDocumentProductLink)
        .where(ContractDocumentProductLink.contract_document_id == str(document.id))
    )

    if document.doc_type != "invoice" or not deal_product_ids:
        await db.commit()
        return

    contract = await Contract.get_by_id(db, str(document.contract_id))
    if not contract or not contract.deal_id:
        raise HTTPException(status_code=400, detail="Invoice product links require contract deal")

    deal_products = await DealProduct.get_by_deal(db, str(contract.deal_id))
    available = {_normalize_link_key(item.id): str(item.id) for item in deal_products}

    links = []
    for requested_id in deal_product_ids:
        real_id = available.get(_normalize_link_key(requested_id))
        if not real_id:
            raise HTTPException(status_code=400, detail="Deal product does not belong to contract deal")
        links.append(ContractDocumentProductLink(
            contract_document_id=str(document.id),
            deal_product_id=real_id,
        ))
    for link in links:
        db.add(link)
    await db.commit()


@router.get("/{contract_id}/documents", response_model=List[ContractDocumentResponse])
async def list_contract_documents(
    contract_id: str,
    doc_type: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
):
    contract = await Contract.get_by_id(db, contract_id)
    if not contract:
        raise HTTPException(status_code=404, detail="Contract not found")
    if doc_type and doc_type not in ALLOWED_DOC_TYPES:
        raise HTTPException(status_code=400, detail="Invalid doc_type")
    documents = await ContractDocument.get_by_contract(db, contract_id, doc_type=doc_type)
    await _attach_document_product_links(db, documents)
    return documents


@router.post("/{contract_id}/documents/upload", response_model=ContractDocumentResponse)
async def upload_contract_document(
    contract_id: str,
    doc_type: str = Form(...),
    file_kind: str = Form(...),
    status: Optional[str] = Form("draft"),
    amount: Optional[float] = Form(None),
    product_ids: Optional[str] = Form(None),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user=Depends(CurrentUser),
):
    if doc_type not in ALLOWED_DOC_TYPES:
        raise HTTPException(status_code=400, detail="Invalid doc_type")
    if file_kind not in ALLOWED_FILE_KINDS:
        raise HTTPException(status_code=400, detail="Invalid file_kind")
    if status and status not in ALLOWED_DOC_STATUSES:
        raise HTTPException(status_code=400, detail="Invalid status")
    contract = await Contract.get_by_id(db, contract_id)
    if not contract:
        raise HTTPException(status_code=404, detail="Contract not found")

    # Serialize MAX(number_in_contract)+1 -> INSERT (no UNIQUE constraint,
    # so a race silently duplicates the per-contract document number).
    async with sequence_lock("contract_document_number"):
        documents = await ContractDocument.get_by_contract_and_type(db, contract_id, doc_type)
        next_number = (max((d.number_in_contract for d in documents), default=0) + 1) if documents else 1

        content = await file.read()
        safe_name = clean_name(file.filename or f"{doc_type}_{next_number}")
        base_path = f"{(settings.STORAGE_LOCAL_ROOT or '').rstrip('/')}/contracts/{contract_id}/{doc_type}"
        await ensure_path(base_path)
        storage_path = f"{base_path}/{doc_type}_{next_number}_{file_kind}_{safe_name}"
        await upload_bytes_with_safe_extension(storage_path, content)

        payload = {
            "contract_id": contract.id,
            "doc_type": doc_type,
            "number_in_contract": next_number,
            "status": status or "draft",
            "amount": _document_amount(amount) if doc_type in {"contract", "addendum", "act"} else None,
        }
        uploaded_by = _user_display_name(user)
        uploaded_at = datetime.now()
        if file_kind == "pdf":
            payload["pdf_file_name"] = file.filename
            payload["pdf_storage_path"] = storage_path
            payload["pdf_uploaded_by"] = uploaded_by
            payload["pdf_uploaded_at"] = uploaded_at
        else:
            payload["edit_file_name"] = file.filename
            payload["edit_storage_path"] = storage_path
            payload["edit_uploaded_by"] = uploaded_by
            payload["edit_uploaded_at"] = uploaded_at

        document = await ContractDocument.create(db, **payload)
    await _set_invoice_product_links(db, document, _parse_product_ids(product_ids))
    await _attach_document_product_links(db, [document])
    if contract.deal_id:
        try:
            await log_event(
                db,
                entity_type="deal",
                entity_id=str(contract.deal_id),
                action="contract.document.upload",
                created_by=str(getattr(user, "id", "")),
                details={
                    "deal_id": str(contract.deal_id),
                    "contract_id": str(contract.id),
                    "contract_number": contract.contract_number,
                    "document_id": str(document.id),
                    "doc_type": document.doc_type,
                    "status": document.status,
                },
            )
        except Exception:
            pass
    return document


@router.post("/documents/{document_id}/upload", response_model=ContractDocumentResponse)
async def upload_contract_document_file(
    document_id: str,
    file_kind: str = Form(...),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user=Depends(CurrentUser),
):
    if file_kind not in ALLOWED_FILE_KINDS:
        raise HTTPException(status_code=400, detail="Invalid file_kind")
    document = await ContractDocument.get_by_id(db, document_id)
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")

    content = await file.read()
    safe_name = clean_name(file.filename or f"{document.doc_type}_{document.number_in_contract}")
    base_path = f"{(settings.STORAGE_LOCAL_ROOT or '').rstrip('/')}/contracts/{document.contract_id}/{document.doc_type}"
    await ensure_path(base_path)
    storage_path = f"{base_path}/{document.doc_type}_{document.number_in_contract}_{file_kind}_{safe_name}"
    await upload_bytes_with_safe_extension(storage_path, content)

    uploaded_by = _user_display_name(user)
    uploaded_at = datetime.now()
    update_payload = {}
    if file_kind == "pdf":
        update_payload["pdf_file_name"] = file.filename
        update_payload["pdf_storage_path"] = storage_path
        update_payload["pdf_uploaded_by"] = uploaded_by
        update_payload["pdf_uploaded_at"] = uploaded_at
    else:
        update_payload["edit_file_name"] = file.filename
        update_payload["edit_storage_path"] = storage_path
        update_payload["edit_uploaded_by"] = uploaded_by
        update_payload["edit_uploaded_at"] = uploaded_at

    updated = await ContractDocument.update(db, document_id, **update_payload)
    if not updated:
        raise HTTPException(status_code=404, detail="Document not found")
    return updated


@router.patch("/documents/{document_id}", response_model=ContractDocumentResponse)
async def update_contract_document(
    document_id: str,
    payload: ContractDocumentUpdate,
    db: AsyncSession = Depends(get_db),
):
    if payload.status and payload.status not in ALLOWED_DOC_STATUSES:
        raise HTTPException(status_code=400, detail="Invalid status")
    current = await ContractDocument.get_by_id(db, document_id)
    if not current:
        raise HTTPException(status_code=404, detail="Document not found")
    update_data = payload.dict(exclude_unset=True)
    if "amount" in update_data:
        update_data["amount"] = _document_amount(update_data.get("amount")) if current.doc_type in {"contract", "addendum", "act"} else None
    document = await ContractDocument.update(db, document_id, **update_data)
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")
    await _attach_document_product_links(db, [document])
    return document


@router.patch("/documents/{document_id}/products", response_model=ContractDocumentResponse)
async def update_contract_document_products(
    document_id: str,
    payload: DocumentProductLinksUpdate,
    db: AsyncSession = Depends(get_db),
):
    document = await ContractDocument.get_by_id(db, document_id)
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")
    if document.doc_type != "invoice":
        raise HTTPException(status_code=400, detail="Product links are available only for invoices")
    await _set_invoice_product_links(db, document, payload.deal_product_ids)
    updated = await ContractDocument.get_by_id(db, document_id)
    await _attach_document_product_links(db, [updated])
    return updated


@router.delete("/documents/{document_id}")
async def delete_contract_document(
    document_id: str,
    db: AsyncSession = Depends(get_db),
):
    document = await ContractDocument.get_by_id(db, document_id)
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")
    await db.execute(
        delete(ContractDocumentProductLink)
        .where(ContractDocumentProductLink.contract_document_id == str(document.id))
    )
    await db.commit()
    success = await ContractDocument.delete(db, document_id)
    if not success:
        raise HTTPException(status_code=404, detail="Document not found")
    await _renumber_documents(db, str(document.contract_id), document.doc_type)
    return {"message": "Document deleted"}


@router.delete("/documents/{document_id}/file", response_model=ContractDocumentResponse)
async def delete_contract_document_file(
    document_id: str,
    file_kind: str = Query(...),
    db: AsyncSession = Depends(get_db),
):
    if file_kind not in ALLOWED_FILE_KINDS:
        raise HTTPException(status_code=400, detail="Invalid file_kind")
    document = await ContractDocument.get_by_id(db, document_id)
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")

    path = document.pdf_storage_path if file_kind == "pdf" else document.edit_storage_path
    if not path:
        raise HTTPException(status_code=404, detail="File not found")

    try:
        await delete_path(path)
    except Exception:
        # If file is already missing in storage, still clear DB pointers.
        pass

    update_payload = {}
    if file_kind == "pdf":
        update_payload["pdf_storage_path"] = None
        update_payload["pdf_file_name"] = None
    else:
        update_payload["edit_storage_path"] = None
        update_payload["edit_file_name"] = None

    updated = await ContractDocument.update(db, document_id, **update_payload)
    if not updated:
        raise HTTPException(status_code=404, detail="Document not found")
    return updated


@router.get("/documents/{document_id}/download")
async def download_contract_document(
    document_id: str,
    file_kind: str = Query("pdf"),
    db: AsyncSession = Depends(get_db),
):
    if file_kind not in ALLOWED_FILE_KINDS:
        raise HTTPException(status_code=400, detail="Invalid file_kind")
    document = await ContractDocument.get_by_id(db, document_id)
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")
    path = document.pdf_storage_path if file_kind == "pdf" else document.edit_storage_path
    if not path:
        raise HTTPException(status_code=404, detail="File not found")
    if not storage_available():
        raise HTTPException(status_code=500, detail="Storage is not configured")
    filename = document.pdf_file_name if file_kind == "pdf" else document.edit_file_name
    if not filename:
        filename = "document.pdf" if file_kind == "pdf" else "document.docx"
    if is_local_storage():
        try:
            content = await read_file_bytes(path)
        except FileNotFoundError:
            raise HTTPException(status_code=404, detail="File not found")
        except Exception:
            raise HTTPException(status_code=502, detail="Failed to download file")
        content_type = mimetypes.guess_type(filename)[0] or "application/octet-stream"
        headers = {"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"}
        return Response(content=content, media_type=content_type, headers=headers)
    try:
        href = await get_download_href(path)
    except httpx.HTTPStatusError as exc:
        status = exc.response.status_code
        if status in {404, 410}:
            raise HTTPException(status_code=404, detail="File not found in storage")
        raise HTTPException(status_code=502, detail="Failed to resolve storage download link")
    except Exception:
        raise HTTPException(status_code=502, detail="Failed to resolve storage download link")

    try:
        async with httpx.AsyncClient(timeout=120, follow_redirects=True) as client:
            response = await client.get(href)
            response.raise_for_status()
    except httpx.HTTPStatusError as exc:
        status = exc.response.status_code
        if status in {404, 410}:
            raise HTTPException(status_code=404, detail="File not found in storage")
        raise HTTPException(status_code=502, detail="Failed to download file from storage")
    except httpx.HTTPError:
        raise HTTPException(status_code=502, detail="Failed to download file from storage")
    except Exception:
        raise HTTPException(status_code=502, detail="Failed to download file from storage")

    content_type = response.headers.get("Content-Type", "application/octet-stream")
    headers = {
        "Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"
    }
    return Response(content=response.content, media_type=content_type, headers=headers)


# ---- Type / status labels for the Excel export (human-readable RU) ----
_CONTRACT_TYPE_LABELS = {
    "expenses": "Расходы",
    "services": "Услуги",
    "labor": "Работы",
    "general_contractor": "Генподряд",
    "partial_contractor": "Подряд (частичный)",
    "subcontractor": "Субподряд",
    "supply_out": "Поставка (исходящая)",
    "supply_in": "Поставка (входящая)",
}
_CONTRACT_STATUS_LABELS = {
    "approval": "На согласовании",
    "in_progress": "В работе",
    "completed": "Завершён",
}


def _doc_last_upload(documents):
    """Return (uploaded_by, uploaded_at) of the most recent file upload across
    all documents of a contract, or (None, None) if no audit data is present."""
    best_at = None
    best_by = None
    for doc in documents:
        for at, by in (
            (getattr(doc, "pdf_uploaded_at", None), getattr(doc, "pdf_uploaded_by", None)),
            (getattr(doc, "edit_uploaded_at", None), getattr(doc, "edit_uploaded_by", None)),
        ):
            if at is None:
                continue
            if best_at is None or at > best_at:
                best_at = at
                best_by = by
    return best_by, best_at


@router.get("/export/xlsx")
async def export_contracts_xlsx(
    request: Request,
    search: Optional[str] = Query(None),
    contract_type: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    customer_id: Optional[str] = Query(None),
    executor_id: Optional[str] = Query(None),
    deal_id: Optional[str] = Query(None),
    sort_by: Optional[str] = Query("contract_date"),
    sort_dir: Optional[str] = Query("desc"),
    db: AsyncSession = Depends(get_db),
    user=Depends(CurrentUser),
):
    """Export the filtered contracts registry to an .xlsx file.

    Accepts the same filter/sort query params as GET /contracts/ but without
    pagination — all matched contracts are exported. Reuses Contract.search_all
    so filtering/permission logic is not duplicated.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Excel export is unavailable: openpyxl is not installed",
        )

    read_all, read_assigned = await get_section_permissions(db, user.role_id, "contracts")
    allowed = None
    if not read_all:
        if not read_assigned:
            items = []
        else:
            allowed = await allowed_deal_ids(db, request, user)
            if allowed == []:
                items = []
            else:
                items = None
    else:
        items = None

    if items is None:
        # Pull every matched contract (no pagination) reusing the registry query.
        items, _total = await Contract.search_all(
            db,
            skip=0,
            limit=10**9,
            search=search,
            contract_type=contract_type,
            status=status,
            customer_id=customer_id,
            executor_id=executor_id,
            deal_id=deal_id,
            sort_by=sort_by,
            sort_dir=sort_dir,
            allowed_deal_ids=allowed,
        )
        await _apply_service_paid_amounts(db, items)

    # Bulk-resolve related names (same fields the registry surfaces).
    company_cache: Dict[str, Optional[str]] = {}
    deal_cache: Dict[str, Optional[str]] = {}

    async def _company_name(company_id) -> Optional[str]:
        if not company_id:
            return None
        key = str(company_id)
        if key not in company_cache:
            company = await Company.get_by_id(db, key)
            company_cache[key] = (
                (company.short_name or company.name) if company else None
            )
        return company_cache[key]

    async def _deal_title(d_id) -> Optional[str]:
        if not d_id:
            return None
        key = str(d_id)
        if key not in deal_cache:
            deal = await Deal.get_by_id(db, key)
            deal_cache[key] = deal.title if deal else None
        return deal_cache[key]

    wb = Workbook()
    ws = wb.active
    ws.title = "Договоры"

    headers = [
        "Номер договора",
        "Дата",
        "Статус",
        "Тип",
        "Сумма",
        "Заказчик",
        "Исполнитель",
        "Сделка",
        "Кол-во документов",
        "Последняя загрузка — кто",
        "Последняя загрузка — когда",
    ]
    ws.append(headers)
    bold = Font(bold=True)
    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=1, column=col_idx).font = bold

    for contract in items:
        documents = await ContractDocument.get_by_contract(db, str(contract.id))
        last_by, last_at = _doc_last_upload(documents)
        amount = float(contract.amount) if contract.amount is not None else None
        ws.append([
            contract.contract_number,
            contract.contract_date.isoformat() if contract.contract_date else None,
            _CONTRACT_STATUS_LABELS.get(contract.status, contract.status),
            _CONTRACT_TYPE_LABELS.get(contract.contract_type, contract.contract_type),
            amount,
            await _company_name(contract.customer_id),
            await _company_name(contract.executor_id),
            await _deal_title(contract.deal_id),
            len(documents),
            last_by,
            last_at.strftime("%Y-%m-%d %H:%M") if last_at else None,
        ])

    # Bold header already set; enable autofilter + sensible column widths.
    last_col_letter = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col_letter}{ws.max_row}"
    widths = [22, 12, 18, 22, 16, 32, 32, 32, 16, 26, 20]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    today = datetime.now().strftime("%Y-%m-%d")
    filename = f"contracts_registry_{today}.xlsx"
    headers_out = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers_out,
    )


@router.get("/{contract_id}/documents/zip")
async def download_contract_documents_zip(
    contract_id: str,
    db: AsyncSession = Depends(get_db),
    user=Depends(CurrentUser),
):
    """Bundle all files (pdf + edit) of every document of a contract into a ZIP.

    Folder structure inside the archive is grouped by doc_type; missing files
    on disk are skipped. Returns 404 if no files exist at all.
    """
    contract = await Contract.get_by_id(db, contract_id)
    if not contract:
        raise HTTPException(status_code=404, detail="Contract not found")
    if not storage_available():
        raise HTTPException(status_code=500, detail="Storage is not configured")

    documents = await ContractDocument.get_by_contract(db, contract_id)

    buffer = io.BytesIO()
    added = 0
    used_names: set = set()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for doc in documents:
            for file_kind, path, file_name in (
                ("pdf", doc.pdf_storage_path, doc.pdf_file_name),
                ("edit", doc.edit_storage_path, doc.edit_file_name),
            ):
                if not path:
                    continue
                try:
                    content = await read_file_bytes(path)
                except FileNotFoundError:
                    continue
                except Exception:
                    continue
                base_name = file_name or f"{doc.doc_type}_{doc.number_in_contract}_{file_kind}"
                arc_name = f"{doc.doc_type}/{doc.number_in_contract}_{clean_name(base_name)}"
                # Avoid in-archive name collisions.
                candidate = arc_name
                suffix = 1
                while candidate in used_names:
                    if "." in arc_name.rsplit("/", 1)[-1]:
                        stem, ext = arc_name.rsplit(".", 1)
                        candidate = f"{stem}_{suffix}.{ext}"
                    else:
                        candidate = f"{arc_name}_{suffix}"
                    suffix += 1
                used_names.add(candidate)
                zf.writestr(candidate, content)
                added += 1

    if added == 0:
        raise HTTPException(status_code=404, detail="No document files found for this contract")

    buffer.seek(0)
    safe_number = clean_name(contract.contract_number) if contract.contract_number else str(contract.id)
    download_name = f"contract_{safe_number}_documents.zip"
    headers_out = {
        "Content-Disposition": f"attachment; filename*=UTF-8''{quote(download_name)}"
    }
    return StreamingResponse(buffer, media_type="application/zip", headers=headers_out)


@router.get("/{contract_id}/card", response_model=ContractCardResponse)
async def get_contract_card(
    contract_id: str,
    db: AsyncSession = Depends(get_db),
):
    contract = await Contract.get_by_id(db, contract_id)
    if not contract:
        raise HTTPException(status_code=404, detail="Contract not found")

    documents = await ContractDocument.get_by_contract(db, contract_id)
    await _attach_document_product_links(db, documents)

    deal_title = None
    subcontractor_title = None
    if contract.deal_id:
        deal = await Deal.get_by_id(db, str(contract.deal_id))
        if deal:
            deal_title = deal.title
    if contract.subcontractor_card_id:
        card = await SubcontractorCard.get_by_id(db, str(contract.subcontractor_card_id))
        if card:
            subcontractor_title = card.title

    contract_id_str = str(contract.id)
    contract_id_hex = contract.id.hex
    entry_result = await db.execute(
        select(IncomeExpenseEntry)
        .where(or_(IncomeExpenseEntry.contract_id == contract_id_str, IncomeExpenseEntry.contract_id == contract_id_hex))
        .order_by(IncomeExpenseEntry.plan_date.desc())
    )
    entries = entry_result.scalars().all()
    payments = await _build_payment_responses(db, entries)

    total_amount = sum(entry.amount for entry in payments) if payments else 0.0
    paid_amount = sum(entry.paid_amount for entry in payments) if payments else 0.0
    pending_amount = max(total_amount - paid_amount, 0.0)
    if contract.contract_type == "services":
        set_committed_value(contract, "amount", float(paid_amount))

    stages_summary: List[ContractStageSummary] = []
    if contract.contract_type == "subcontractor":
        contract_matches = _id_conditions(SubcontractorStage.contract_id, str(contract.id))
        stage_result = await db.execute(select(SubcontractorStage).where(contract_matches))
        stages = stage_result.scalars().all()
        stages = [s for s in stages if (s.planned_cost or 0) > 0 and s.stage_type == "stage"]
        for stage in stages:
            stages_summary.append(ContractStageSummary(
                id=str(stage.id),
                name=stage.name,
                stage_type=stage.stage_type,
                planned_cost=stage.planned_cost or 0.0,
                date_start=stage.date_start,
                date_end=stage.date_end,
                status=stage.status,
                is_closed=None,
            ))
    elif contract.deal_id:
        stages = await Stage.get_by_deal_id(db, str(contract.deal_id))
        stages = [s for s in stages if (s.planned_cost or 0) > 0 and s.stage_type == "stage"]
        stage_payment_maps = await _build_stage_payment_maps(db, [str(s.id) for s in stages])
        total_map = stage_payment_maps["total"]
        paid_map = stage_payment_maps["paid"]
        for stage in stages:
            stage_id = str(stage.id)
            stage_id_hex = stage.id.hex
            total = total_map.get(stage_id, total_map.get(stage_id_hex, 0.0))
            paid = paid_map.get(stage_id, paid_map.get(stage_id_hex, 0.0))
            is_closed = stage.is_closed
            if stage.stage_type == "payment" and total > 0:
                is_closed = paid + 1e-9 >= total
            stages_summary.append(ContractStageSummary(
                id=stage_id,
                name=stage.name,
                stage_type=stage.stage_type,
                planned_cost=stage.planned_cost or 0.0,
                date_start=stage.date_start,
                date_end=stage.date_end,
                status=stage.status,
                is_closed=is_closed,
            ))

    return ContractCardResponse(
        contract=contract,
        deal_title=deal_title,
        subcontractor_title=subcontractor_title,
        documents=documents,
        payment_summary=ContractPaymentSummary(
            total_amount=total_amount,
            paid_amount=paid_amount,
            pending_amount=pending_amount,
        ),
        payments=payments,
        stages=stages_summary,
    )


@router.get("/{contract_id}/expenses", response_model=ContractExpenseCardResponse)
async def get_contract_expenses(
    contract_id: str,
    db: AsyncSession = Depends(get_db),
):
    contract = await _get_service_contract_or_400(db, contract_id)
    result = await db.execute(
        select(IncomeExpenseEntry)
        .where(_contract_id_conditions(contract_id))
        .where(IncomeExpenseEntry.direction == "expense")
        .order_by(IncomeExpenseEntry.plan_date.desc())
    )
    entries = result.scalars().all()
    payments = await _build_payment_responses(db, entries)
    total_amount = sum(entry.amount for entry in payments) if payments else 0.0
    paid_amount = sum(entry.paid_amount for entry in payments) if payments else 0.0
    pending_amount = max(total_amount - paid_amount, 0.0)
    return ContractExpenseCardResponse(
        summary=ContractPaymentSummary(
            total_amount=total_amount,
            paid_amount=paid_amount,
            pending_amount=pending_amount,
        ),
        items=payments,
    )


@router.post("/{contract_id}/expenses", response_model=IncomeExpenseEntryResponse)
async def create_contract_expense(
    contract_id: str,
    payload: ContractExpenseCreate,
    db: AsyncSession = Depends(get_db),
):
    contract = await _get_service_contract_or_400(db, contract_id)
    entry = IncomeExpenseEntry(
        direction="expense",
        amount=payload.amount,
        plan_date=payload.plan_date,
        actual_date=payload.actual_date,
        payer_id=str(contract.customer_id) if contract.customer_id else None,
        payee_id=str(contract.executor_id) if contract.executor_id else None,
        deal_id=str(contract.deal_id) if contract.deal_id else None,
        contract_id=str(contract.id),
        category_code=payload.category_code,
    )
    db.add(entry)
    await db.commit()
    await db.refresh(entry)
    responses = await _build_payment_responses(db, [entry])
    return responses[0]


@router.post("/{contract_id}/expenses/bulk", response_model=List[IncomeExpenseEntryResponse])
async def create_contract_expenses_bulk(
    contract_id: str,
    payload: ContractExpenseBulkCreate,
    db: AsyncSession = Depends(get_db),
):
    contract = await _get_service_contract_or_400(db, contract_id)
    if payload.frequency not in ALLOWED_EXPENSE_FREQUENCIES:
        raise HTTPException(status_code=400, detail="Invalid frequency")
    if payload.periods <= 0:
        raise HTTPException(status_code=400, detail="Periods must be positive")
    entries: List[IncomeExpenseEntry] = []
    anchor_day = payload.start_date.day
    anchor_is_eom = payload.start_date.day == monthrange(payload.start_date.year, payload.start_date.month)[1]
    current = payload.start_date
    for _ in range(payload.periods):
        entry = IncomeExpenseEntry(
            direction="expense",
            amount=payload.amount,
            plan_date=current,
            actual_date=None,
            payer_id=str(contract.customer_id) if contract.customer_id else None,
            payee_id=str(contract.executor_id) if contract.executor_id else None,
            deal_id=str(contract.deal_id) if contract.deal_id else None,
            contract_id=str(contract.id),
            category_code=payload.category_code,
        )
        db.add(entry)
        entries.append(entry)
        if payload.frequency == "week":
            current = _next_period_date(current, payload.frequency)
        else:
            if payload.frequency == "month":
                months_step = 1
            elif payload.frequency == "quarter":
                months_step = 3
            else:
                months_step = 12
            current = _add_months_with_anchor(current, months_step, anchor_day, anchor_is_eom)
    await db.commit()
    for entry in entries:
        await db.refresh(entry)
    return await _build_payment_responses(db, entries)


@router.patch("/{contract_id}/expenses/{entry_id}", response_model=IncomeExpenseEntryResponse)
async def update_contract_expense(
    contract_id: str,
    entry_id: str,
    payload: ContractExpenseUpdate,
    db: AsyncSession = Depends(get_db),
):
    await _get_service_contract_or_400(db, contract_id)
    result = await db.execute(select(IncomeExpenseEntry).where(IncomeExpenseEntry.id == entry_id))
    entry = result.scalar_one_or_none()
    if not entry or not entry.contract_id:
        raise HTTPException(status_code=404, detail="Entry not found")
    if str(entry.contract_id) not in {contract_id, contract_id.replace("-", "")}:
        raise HTTPException(status_code=400, detail="Entry does not belong to contract")
    if entry.direction != "expense":
        raise HTTPException(status_code=400, detail="Only expense entries allowed")
    update_data = payload.dict(exclude_unset=True)
    for key, value in update_data.items():
        setattr(entry, key, value)
    await db.commit()
    await db.refresh(entry)
    responses = await _build_payment_responses(db, [entry])
    return responses[0]


@router.delete("/{contract_id}/expenses/{entry_id}")
async def delete_contract_expense(
    contract_id: str,
    entry_id: str,
    db: AsyncSession = Depends(get_db),
):
    await _get_service_contract_or_400(db, contract_id)
    result = await db.execute(select(IncomeExpenseEntry).where(IncomeExpenseEntry.id == entry_id))
    entry = result.scalar_one_or_none()
    if not entry or not entry.contract_id:
        raise HTTPException(status_code=404, detail="Entry not found")
    if str(entry.contract_id) not in {contract_id, contract_id.replace("-", "")}:
        raise HTTPException(status_code=400, detail="Entry does not belong to contract")
    if entry.direction != "expense":
        raise HTTPException(status_code=400, detail="Only expense entries allowed")
    await db.delete(entry)
    await db.commit()
    return {"message": "Deleted"}
